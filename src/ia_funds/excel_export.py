from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from ia_funds.loader import wide_to_long

log = logging.getLogger(__name__)


def export_workbook(wide: pd.DataFrame, dest: str | Path) -> Path:
    """
    Write an Excel workbook with:
    - Wide: original matrix (Funds, Asset class, Code, dates)
    - Long: melted NAV history
    - Summary: last available NAV and simple change vs prior column
    """
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    log.info("Building Excel workbook: %s", dest)
    long = wide_to_long(wide)

    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Wide"
    for row in dataframe_to_rows(wide, index=False, header=True):
        ws0.append(row)

    ws1 = wb.create_sheet("Long")
    for row in dataframe_to_rows(long, index=False, header=True):
        ws1.append(row)

    meta = ["Funds", "Asset class", "Code"]
    date_cols = [c for c in wide.columns if c not in meta]
    if len(date_cols) < 2:
        summary = wide[meta].copy()
        summary["last_date"] = date_cols[0] if date_cols else ""
        summary["last_nav"] = wide[date_cols[0]] if date_cols else pd.NA
        summary["chg_vs_prior"] = pd.NA
    else:
        d_last, d_prev = date_cols[-1], date_cols[-2]
        summary = wide[meta].copy()
        summary["last_date"] = d_last
        summary["last_nav"] = wide[d_last]
        prev = wide[d_prev]
        summary["chg_vs_prior"] = (wide[d_last] / prev - 1.0).where(prev.notna() & (prev != 0))

    ws2 = wb.create_sheet("Summary")
    for row in dataframe_to_rows(summary, index=False, header=True):
        ws2.append(row)

    wb.save(dest)
    log.info("Saved Excel workbook (%d wide rows, %d long rows)", len(wide), len(long))
    return dest
